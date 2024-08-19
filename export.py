import openpyxl
from openpyxl.utils import get_column_letter
import income
import expenses
import budget
import savings
import os

def export_to_excel(file_name="personal_finance_report.xlsx"):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Personal Finance Report"

        # Write headers
        headers = ["Category", "Description", "Amount", "Type"]
        for i, header in enumerate(headers, start=1):
            sheet[f"{get_column_letter(i)}1"] = header

        # Populate data
        row = 2
        row = populate_sheet(sheet, row, "Income", income.load_data())
        row = populate_sheet(sheet, row, "Expense", expenses.load_data(), include_category=True)
        row = populate_sheet(sheet, row, "Budget", budget.load_data().items(), budget=True)
        populate_savings(sheet, row, savings.load_data())

        # Adjust column widths
        adjust_column_width(sheet)
        workbook.save(file_name)
        print(f"Data successfully exported to '{file_name}'.")

        # Clear JSON files
        clear_data_files()
    except Exception as e:
        print(f"An error occurred while exporting data: {e}")

def populate_sheet(sheet, start_row, record_type, data, include_category=False, budget=False):
    for i, entry in enumerate(data, start=start_row):
        if budget:
            category, amount = entry
            sheet[f"A{i}"] = category
            sheet[f"C{i}"] = amount
        else:
            if include_category:
                sheet[f"A{i}"] = entry['category']
            sheet[f"B{i}"] = entry.get('source', entry.get('description', ''))
            sheet[f"C{i}"] = entry['amount']

        sheet[f"D{i}"] = record_type

    return start_row + len(data)

def populate_savings(sheet, start_row, data):
    sheet[f"A{start_row}"] = "Savings"
    sheet[f"B{start_row}"] = "Goal"
    sheet[f"C{start_row}"] = data["goal"]

    sheet[f"A{start_row + 1}"] = "Savings"
    sheet[f"B{start_row + 1}"] = "Saved"
    sheet[f"C{start_row + 1}"] = data["saved"]

def adjust_column_width(sheet):
    for col in sheet.columns:
        max_length = max(len(str(cell.value) or '') for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max(max_length + 2, 10)

def clear_data_files():
    files = ['data/income.json', 'data/expenses.json', 'data/budget.json', 'data/savings.json']
    for file in files:
        if os.path.exists(file):
            with open(file, 'w') as f:
                f.write('[]' if 'budget' not in file else '{}')
    print("All data files have been cleared.")
